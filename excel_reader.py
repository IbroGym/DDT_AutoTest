from __future__ import annotations
from dataclasses import dataclass
from typing import List
import openpyxl


@dataclass(frozen=True)
class LoginCase:
    case_id: str
    username: str
    password: str
    expected: str


def load_login_cases(xlsx_path: str, sheet_name: str = "Sheet1") -> List[LoginCase]:
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header row

    cases: List[LoginCase] = []
    for r in rows:
        case_id, username, password, expected = r

        # basic validation (helps catch empty cells early)
        if not all([case_id, username, password, expected]):
            raise ValueError(f"Empty value in row: {r}")

        cases.append(
            LoginCase(
                case_id=str(case_id).strip(),
                username=str(username).strip(),
                password=str(password).strip(),
                expected=str(expected).strip(),
            )
        )

    return cases
